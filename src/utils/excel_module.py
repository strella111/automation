import datetime
import json

from openpyxl import load_workbook, Workbook

import csv
from pathlib import Path
import os
import numpy as np

try:
    from PyQt5 import QtCore
    from config.settings_manager import get_main_settings
except Exception:
    QtCore = None
    get_main_settings = None

from typing import List, Optional
from loguru import logger
from core.common.enums import Channel, Direction


class CalibrationCSV:
    """Класс для работы с CSV файлами калибровки фазировки"""

    def __init__(self, bu_address: int):
        """
        Инициализация для работы с CSV файлом калибровки

        Args:
            bu_address: Адрес БУ для формирования имени файла
        """
        self.bu_address = bu_address
        # Определяем базовую директорию сохранения для режима phase
        base_dir = None
        try:
            if get_main_settings is not None:
                qsettings = get_main_settings()
                v = qsettings.value('base_save_dir')
                if v:
                    base_dir = str(v)
        except Exception:
            base_dir = None

        if base_dir:
            self.calbs_dir = Path(os.path.join(base_dir, 'phase'))
        else:
            self.calbs_dir = Path("calbs")
        self.csv_file = self.calbs_dir / f"{bu_address}.csv"

        self.calbs_dir.mkdir(exist_ok=True)

        self.columns = [
            "Передатчик_Горизонтальная",
            "Передатчик_Вертикальная",
            "Приемник_Вертикальная",
            "Приемник_Горизонтальная"
        ]

        self._initialize_csv_if_needed()

    def _initialize_csv_if_needed(self):
        """Создает CSV файл с нулевыми значениями если файл не существует
        
        Структура файла:
        - 16 блоков (ЛЗ 0-15)
        - Каждый блок: 32 ППМ × 4 столбца = 128 строк
        """
        if not self.csv_file.exists():
            logger.info(f"Создание нового `CSV` файла: {self.csv_file}")

            data = []
            # 16 блоков × 32 ППМ
            for block in range(16):  # ЛЗ 0-15
                for ppm in range(32):  # 32 ППМ в каждом блоке
                    row = [0, 0, 0, 0]
                    data.append(row)

            with open(self.csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerows(data)

            logger.info(f"CSV файл инициализирован с {len(data)} строками (17 блоков × 32 ППМ)")

    def get_column_index(self, channel: Channel, direction: Direction) -> int:
        """
        Возвращает индекс столбца для заданной комбинации канал/поляризация

        Args:
            channel: Канал (передатчик/приемник)
            direction: Поляризация (вертикальная/горизонтальная)

        Returns:
            int: Индекс столбца (0-3)
        """
        channel_name = "Передатчик" if channel == Channel.Transmitter else "Приемник"
        direction_name = "Вертикальная" if direction == Direction.Vertical else "Горизонтальная"
        column_name = f"{channel_name}_{direction_name}"

        try:
            return self.columns.index(column_name)
        except ValueError:
            logger.error(f"Неизвестная комбинация канал/поляризация: {column_name}")
            return 0

    def save_phase_results(self, channel: Channel, direction: Direction, phase_results: List[int], 
                           delay_line_discretes: Optional[List[int]] = None):
        """
        Сохраняет результаты фазировки в соответствующий столбец CSV файла
        
        Файл содержит 16 блоков:
        - Блок 0 (ЛЗ=0): базовая калибровка ФВ
        - Блоки 1-15 (ЛЗ=1-15): базовая калибровка + дискреты ЛЗ

        Args:
            channel: Канал (передатчик/приемник)
            direction: Поляризация (вертикальная/горизонтальная)
            phase_results: Список дискретов фазовращателей для 32 ППМ (базовая калибровка)
            delay_line_discretes: Список из 16 значений дискретов для ЛЗ 0-15
        """
        if len(phase_results) != 32:
            logger.error(f"Ожидается 32 значения, получено {len(phase_results)}")
            return

        column_index = self.get_column_index(channel, direction)
        column_name = self.columns[column_index]

        logger.info(f"Сохранение результатов фазировки в столбец '{column_name}' файла {self.csv_file}")
        if delay_line_discretes:
            logger.info(f"Фазировка ЛЗ включена, сохранение 16 блоков с коррекцией ЛЗ")
        else:
            logger.info(f"Фазировка ЛЗ выключена, копирование базовой калибровки во все блоки")

        try:
            existing_data = []
            try:
                with open(self.csv_file, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f, delimiter=';')
                    for row in reader:
                        int_row = []
                        for i in range(4):
                            if i < len(row):
                                try:
                                    int_row.append(int(row[i]))
                                except ValueError:
                                    int_row.append(0)
                            else:
                                int_row.append(0)
                        existing_data.append(int_row)
            except FileNotFoundError:
                pass


            for ppm_index in range(32):
                row_index = ppm_index
                existing_data[row_index][column_index] = phase_results[ppm_index]

            if delay_line_discretes and len(delay_line_discretes) == 16:
                for block in range(1, 16):
                    for ppm_index in range(32):
                        row_index = block * 32 + ppm_index
                        
                        if block <= 15:
                            value = phase_results[ppm_index] + delay_line_discretes[block]
                            if value > 63:
                                value -= 64
                            existing_data[row_index][column_index] = value


            with open(self.csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerows(existing_data)

            logger.info(f"Результаты фазировки успешно сохранены в столбец '{column_name}' (17 блоков)")

        except Exception as e:
            logger.error(f"Ошибка при сохранении результатов фазировки: {e}")

    def load_phase_results(self, channel: Channel, direction: Direction) -> Optional[List[int]]:
        """
        Загружает результаты фазировки из соответствующего столбца CSV файла

        Args:
            channel: Канал (передатчик/приемник)
            direction: Поляризация (вертикальная/горизонтальная)

        Returns:
            List[int]: Список дискретов фазовращателей для 32 ППМ или None при ошибке
        """
        column_index = self.get_column_index(channel, direction)
        column_name = self.columns[column_index]

        try:
            with open(self.csv_file, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)

                phase_results = []
                for row_index, row in enumerate(reader):
                    if row_index >= 32:  # Только первые 32 ППМ
                        break

                    if column_index < len(row):
                        try:
                            phase_results.append(int(row[column_index]))
                        except ValueError:
                            phase_results.append(0)
                    else:
                        phase_results.append(0)

                while len(phase_results) < 32:
                    phase_results.append(0)

                logger.info(f"Загружены результаты фазировки из столбца '{column_name}': {len(phase_results)} значений")
                return phase_results

        except Exception as e:
            logger.error(f"Ошибка при загрузке результатов фазировки: {e}")
            return None

    def get_file_path(self) -> Path:
        """Возвращает путь к CSV файлу"""
        return self.csv_file

def get_or_create_excel_for_check(base_dir, dir_name, file_name, mode, chanel, direction, spacing=True):
    """
    Проверяет существование Excel файла.
    Если файл существует - открывает и возвращает его.
    Если файл не существует - создает новый и возвращает его.

    Args:
        file_path (str): Путь к Excel файлу

    Returns:
        openpyxl.Workbook: Объект рабочей книги Excel
    """
    try:

        eff_dir = os.path.join(base_dir, dir_name)

        if not os.path.exists(eff_dir):
            os.makedirs(eff_dir)

        file_path = os.path.join(eff_dir, file_name)

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)

        else:
            workbook = Workbook()
            workbook.save(file_path)

        if mode == 'check':
            sheet_name = ''
            if chanel == Channel.Receiver and direction == Direction.Horizontal:
                sheet_name = 'ПРМГ'
            elif chanel == Channel.Receiver and direction == Direction.Vertical:
                sheet_name = 'ПРМВ'
            elif chanel == Channel.Transmitter and direction == Direction.Horizontal:
                sheet_name = 'ПРДГ'
            elif chanel == Channel.Transmitter and direction == Direction.Vertical:
                sheet_name = 'ПРДВ'
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.create_sheet(sheet_name)

            if spacing:
                worksheet.insert_rows(idx=1, amount=42)
            worksheet.cell(1, 1).value = 'DateTime'
            worksheet.cell(1, 2).value = datetime.datetime.now().strftime('%d.%m.%Y %H:%M')
            row = ["Номер ППМ",
                   "Статус",
                   "Амплитуда_abs",
                   "Амплитуда_delta",
                   "Фаза",
                   "Дельта ФВ",
                   "Факт. значение 5.625",
                   "Факт. значение 11.25",
                   "Факт. значение 22.5",
                   "Факт. значение 45",
                   "Факт. значение 90",
                   "Факт. значение 180"]
            for i, value in enumerate(row):
                worksheet.cell(row=2, column=i + 1).value = value

            return worksheet, workbook, file_path
        elif mode == 'stend':
            sheet_name = ''
            if chanel == Channel.Receiver and direction == Direction.Horizontal:
                sheet_name = 'ПРМГ'
            elif chanel == Channel.Receiver and direction == Direction.Vertical:
                sheet_name = 'ПРМВ'
            elif chanel == Channel.Transmitter and direction == Direction.Horizontal:
                sheet_name = 'ПРДГ'
            elif chanel == Channel.Transmitter and direction == Direction.Vertical:
                sheet_name = 'ПРДВ'
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.create_sheet(sheet_name)

            if spacing:
                worksheet.insert_rows(idx=1, amount=42)
            worksheet.cell(1, 1).value = 'DateTime'
            worksheet.cell(1, 2).value = datetime.datetime.now().strftime('%d.%m.%Y %H:%M')
            row = ["Номер ППМ",
                   "0 Амп.",
                   "0 Фаза",
                   "5.625 Амп.",
                   "5.625 Фаза",
                   "11.25 Амп.",
                   "11.25 Фаза",
                   "22.5 Амп.",
                   "22.5 Фаза",
                   "45 Амп.",
                   "45 Фаза",
                   "90 Амп.",
                   "90 Фаза",
                   "180 Амп.",
                   "180 Фаза"]
            for i, value in enumerate(row):
                worksheet.cell(row=2, column=i + 1).value = value

            return worksheet, workbook, file_path

        worksheet = workbook.active
        return worksheet, workbook, file_path


    except Exception as e:
        print(f"Ошибка при работе с файлом {file_path}: {e}")
        return None


def save_beam_pattern_results(base_dir: str, beams: List[int], freq_list: List[float], 
                               data: dict, x_list: List[float], y_list: List[float], 
                               step_x: float, step_y: float, save_dir: Optional[str] = None,
                               scan_params: Optional[dict] = None,
                               pna_settings: Optional[dict] = None,
                               sync_settings: Optional[dict] = None):
    """
    Сохраняет результаты измерения лучей в Excel файлы (формат как в luchi.py)
    
    Структура файла:
    - Для каждого луча отдельный файл: Beam№{beam_num}.xlsx
    - В каждом файле для каждой частоты:
      * Строка 1: 'Frequency' | значение частоты
      * Строка 2: 'Magnitude'
      * Строка 3 до 3+len_x: данные амплитуды (по столбцам - y координаты)
      * Строка 3+len_x+1: 'Phase'
      * Строка 3+len_x+2 до 3+len_x*2+1: данные фазы
    
    Сохраняет в base_dir/luchi/{папка_с_датой}/
    При первом сохранении создается папка с датой, при последующих используется та же папка.
    
    Args:
        base_dir: Базовая директория для сохранения
        beams: Список номеров лучей
        freq_list: Список частот в МГц
        data: Данные измерений {beam_num: {freq: {'x': [...], 'y': [...], 'amp': [[...]], 'phase': [[...]]}}}
        x_list: Список координат X
        y_list: Список координат Y
        step_x: Шаг по X
        step_y: Шаг по Y
        save_dir: Опциональный путь к папке с датой (если передан, используется существующая папка)
    """
    try:
        from openpyxl import Workbook, load_workbook
        
        len_x = len(x_list)
        size_freq_data = 3 + len_x * 2  # Размер блока для одной частоты
        
        # Определяем папку для сохранения
        if save_dir:
            # Используем существующую папку
            final_save_dir = save_dir
            logger.debug(f"Использование существующей папки: {final_save_dir}")
        else:
            # Создаем новую папку с датой
            now_datetime = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            luchi_base_dir = os.path.join(base_dir, 'beams/scan_beams')
            final_save_dir = os.path.join(luchi_base_dir, now_datetime)
            logger.info(f"Создание новой папки: {final_save_dir}")
        
        # Создаем директорию (если не существует)
        os.makedirs(final_save_dir, exist_ok=True)
        
        logger.info(f"Сохранение результатов измерения лучей в {final_save_dir}")
        
        for beam_num in beams:
            if beam_num not in data:
                logger.warning(f"Нет данных для луча {beam_num}, пропускаем")
                continue
            
            file_name = os.path.join(final_save_dir, f"Beam№{beam_num}.xlsx")
            
            # Ищем существующий файл в папке с датой
            existing_file = None
            if os.path.exists(file_name):
                existing_file = file_name
            
            # Открываем существующий файл или создаем новый
            if existing_file:
                workbook = load_workbook(existing_file)
                logger.debug(f"Открыт существующий файл для обновления: {file_name}")
            else:
                workbook = Workbook()
                logger.debug(f"Создан новый файл для луча {beam_num}")
            
            sheet = workbook.active
            
            # Заполняем заголовки и данные для каждой частоты
            for freq_idx, freq in enumerate(freq_list):
                if freq not in data[beam_num]:
                    continue
                
                row_start = freq_idx * size_freq_data + 1
                
                # Frequency
                sheet.cell(row=row_start, column=1).value = 'Frequency'
                sheet.cell(row=row_start, column=2).value = freq
                
                # Magnitude
                sheet.cell(row=row_start + 1, column=1).value = 'Magnitude'
                
                # Phase
                sheet.cell(row=row_start + 2 + len_x, column=1).value = 'Phase'
                
                # Заполняем данные для этой частоты
                freq_data = data[beam_num][freq]
                amp_2d = np.array(freq_data['amp'])
                phase_2d = np.array(freq_data['phase'])
                
                # Записываем амплитуду (строки 3 до 3+len_x-1, столбцы - y координаты)
                # В luchi.py: x - по строкам, y - по столбцам
                for x_idx in range(len_x):
                    for y_idx in range(len(y_list)):
                        if not np.isnan(amp_2d[y_idx, x_idx]):
                            sheet.cell(row=row_start + 2 + x_idx, column=y_idx + 1).value = amp_2d[y_idx, x_idx]
                
                # Записываем фазу (строки 3+len_x+1 до 3+len_x*2, столбцы - y координаты)
                for x_idx in range(len_x):
                    for y_idx in range(len(y_list)):
                        if not np.isnan(phase_2d[y_idx, x_idx]):
                            sheet.cell(row=row_start + 3 + len_x + x_idx, column=y_idx + 1).value = phase_2d[y_idx, x_idx]
            
            # Сохраняем файл (обновляем существующий или создаем новый)
            workbook.save(file_name)
            if existing_file:
                logger.debug(f"Обновлен существующий файл: {file_name}")
            else:
                logger.debug(f"Создан новый файл: {file_name}")
            
            logger.info(f"Файл для луча {beam_num} сохранен: {file_name}")
        
        # Сохраняем параметры сканирования в JSON файл
        if scan_params:
            params_file = os.path.join(final_save_dir, 'scan_params.json')
            try:
                params_data = {
                    'beams': beams,
                    'freq_list': freq_list,
                    'x_list': x_list,
                    'y_list': y_list,
                    'step_x': float(step_x),
                    'step_y': float(step_y),
                    'left_x': float(scan_params.get('left_x', x_list[0] if x_list else 0.0)),
                    'right_x': float(scan_params.get('right_x', x_list[-1] if x_list else 0.0)),
                    'up_y': float(scan_params.get('up_y', y_list[0] if y_list else 0.0)),
                    'down_y': float(scan_params.get('down_y', y_list[-1] if y_list else 0.0)),
                }
                
                # Добавляем настройки PNA
                if pna_settings:
                    # Конвертируем значения в сериализуемый формат
                    pna_params = {}
                    for key, value in pna_settings.items():
                        if isinstance(value, (int, float, str, bool)):
                            pna_params[key] = value
                        elif value is None:
                            pna_params[key] = None
                        else:
                            pna_params[key] = str(value)
                    params_data['pna_settings'] = pna_params
                
                # Добавляем настройки синхронизатора
                if sync_settings:
                    # Конвертируем значения в сериализуемый формат
                    sync_params = {}
                    for key, value in sync_settings.items():
                        if isinstance(value, (int, float, str, bool)):
                            sync_params[key] = value
                        elif value is None:
                            sync_params[key] = None
                        else:
                            sync_params[key] = str(value)
                    params_data['sync_settings'] = sync_params
                
                # Сохраняем время начала измерения (если передано)
                if scan_params and 'measurement_start_time' in scan_params:
                    params_data['measurement_start_time'] = float(scan_params['measurement_start_time'])
                
                with open(params_file, 'w', encoding='utf-8') as f:
                    json.dump(params_data, f, indent=2, ensure_ascii=False)
                logger.debug(f"Параметры сканирования сохранены в {params_file}")
            except Exception as e:
                logger.warning(f"Не удалось сохранить параметры сканирования: {e}")
        
        logger.info(f"Все файлы сохранены в {final_save_dir}")
        return final_save_dir
        
    except Exception as e:
        logger.error(f"Ошибка при сохранении результатов измерения лучей: {e}", exc_info=True)
        return None


def load_beam_pattern_results(save_dir: str) -> Optional[dict]:
    """
    Загружает результаты измерения лучей из Excel файлов для досканирования
    
    Args:
        save_dir: Путь к папке с результатами (base_dir/luchi/{дата})
        
    Returns:
        dict: {
            'beams': [список лучей],
            'freq_list': [список частот],
            'data': {beam_num: {freq: {'x': [...], 'y': [...], 'amp': [[...]], 'phase': [[...]]}}},
            'x_list': [список координат X],
            'y_list': [список координат Y],
            'step_x': шаг по X,
            'step_y': шаг по Y
        } или None при ошибке
    """
    try:
        from openpyxl import load_workbook
        
        if not os.path.exists(save_dir):
            logger.error(f"Папка не найдена: {save_dir}")
            return None

        params_file = os.path.join(save_dir, 'scan_params.json')
        loaded_params = None
        if os.path.exists(params_file):
            try:
                with open(params_file, 'r', encoding='utf-8') as f:
                    loaded_params = json.load(f)
                logger.info(f"Загружены параметры сканирования из {params_file}")
            except Exception as e:
                logger.warning(f"Не удалось загрузить параметры сканирования: {e}")
        
        # Находим все файлы Beam№*.xlsx
        beam_files = []
        for filename in os.listdir(save_dir):
            if filename.startswith('Beam№') and filename.endswith('.xlsx'):
                try:
                    beam_num = int(filename.replace('Beam№', '').replace('.xlsx', ''))
                    beam_files.append((beam_num, os.path.join(save_dir, filename)))
                except ValueError:
                    continue
        
        if not beam_files:
            logger.warning(f"Не найдено файлов лучей в {save_dir}")
            return None
        
        beam_files.sort(key=lambda x: x[0])  # Сортируем по номеру луча
        
        # Используем параметры из JSON, если есть, иначе определяем из файлов
        if loaded_params:
            beams = loaded_params.get('beams', [beam_num for beam_num, _ in beam_files])
            freq_list = loaded_params.get('freq_list', [])
            x_list = loaded_params.get('x_list', [])
            y_list = loaded_params.get('y_list', [])
            step_x = loaded_params.get('step_x', 1.0)
            step_y = loaded_params.get('step_y', 1.0)
        else:
            beams = [beam_num for beam_num, _ in beam_files]
            freq_list = []
            x_list = []
            y_list = []
            step_x = 1.0
            step_y = 1.0
        
        # Загружаем данные из первого файла для определения структуры
        first_beam_num, first_file = beam_files[0]
        workbook = load_workbook(first_file)
        sheet = workbook.active
        
        # Если частоты не загружены из JSON, определяем из файла
        if not freq_list:
            # Ищем все частоты
            row = 1
            while row <= sheet.max_row:
                cell = sheet.cell(row, 1)
                if cell.value == 'Frequency':
                    freq_cell = sheet.cell(row, 2)
                    if freq_cell.value:
                        try:
                            freq = float(freq_cell.value)
                            freq_list.append(freq)
                        except (ValueError, TypeError):
                            pass
                row += 1
            
            if not freq_list:
                logger.error("Не найдено частот в файле")
                return None
        
        # Определяем размеры данных из первой частоты
        first_freq_row = None
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row, 1).value == 'Frequency' and sheet.cell(row, 2).value == freq_list[0]:
                first_freq_row = row
                break
        
        if first_freq_row is None:
            logger.error("Не найдена первая частота в файле")
            return None
        
        # Находим размер данных (количество столбцов с данными)
        magnitude_row = first_freq_row + 1
        max_col = 0
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(magnitude_row + 1, col)  # Первая строка данных
            if cell.value is not None:
                max_col = max(max_col, col)
        
        len_y = max_col  # Количество столбцов = количество Y координат
        
        # Находим количество строк данных (до следующего Frequency или до конца)
        next_freq_row = None
        for row in range(first_freq_row + 1, sheet.max_row + 1):
            if sheet.cell(row, 1).value == 'Frequency':
                next_freq_row = row
                break
        
        if next_freq_row:
            # size_freq_data = 3 + len_x * 2
            # next_freq_row = first_freq_row + size_freq_data
            # len_x = (next_freq_row - first_freq_row - 3) // 2
            len_x = (next_freq_row - first_freq_row - 3) // 2
        else:
            # Последняя частота - считаем до конца
            phase_start_row = None
            for row in range(first_freq_row + 1, sheet.max_row + 1):
                if sheet.cell(row, 1).value == 'Phase':
                    phase_start_row = row
                    break
            if phase_start_row:
                # phase_start_row = first_freq_row + 3 + len_x
                len_x = phase_start_row - first_freq_row - 3
            else:
                # Если не нашли Phase, используем общую формулу
                len_x = (sheet.max_row - first_freq_row - 3) // 2
        
        size_freq_data = 3 + len_x * 2
        
        # Если координаты не загружены из JSON, используем индексы
        if not x_list:
            x_list = list(range(len_x))
        if not y_list:
            y_list = list(range(len_y))
        
        # Загружаем данные для всех лучей
        data = {}
        
        for beam_num, file_path in beam_files:
            data[beam_num] = {}
            
            workbook = load_workbook(file_path)
            sheet = workbook.active
            
            for freq_idx, freq in enumerate(freq_list):
                row_start = freq_idx * size_freq_data + 1
                
                # Инициализируем массивы
                amp_2d = np.full((len_y, len_x), np.nan)
                phase_2d = np.full((len_y, len_x), np.nan)
                
                # Загружаем амплитуду
                for x_idx in range(len_x):
                    for y_idx in range(len_y):
                        cell = sheet.cell(row_start + 2 + x_idx, y_idx + 1)
                        if cell.value is not None:
                            try:
                                amp_2d[y_idx, x_idx] = float(cell.value)
                            except (ValueError, TypeError):
                                pass
                
                # Загружаем фазу
                for x_idx in range(len_x):
                    for y_idx in range(len_y):
                        cell = sheet.cell(row_start + 3 + len_x + x_idx, y_idx + 1)
                        if cell.value is not None:
                            try:
                                phase_2d[y_idx, x_idx] = float(cell.value)
                            except (ValueError, TypeError):
                                pass
                
                data[beam_num][freq] = {
                    'x': x_list,
                    'y': y_list,
                    'amp': amp_2d.tolist(),
                    'phase': phase_2d.tolist()
                }
        
        logger.info(f"Загружены данные: {len(beams)} лучей, {len(freq_list)} частот, {len_x}x{len_y} точек")
        
        result = {
            'beams': beams,
            'freq_list': freq_list,
            'data': data,
            'x_list': x_list,
            'y_list': y_list,
            'step_x': step_x,
            'step_y': step_y,
            'save_dir': save_dir
        }
        
        # Добавляем параметры из JSON, если они были загружены
        if loaded_params:
            result['left_x'] = loaded_params.get('left_x')
            result['right_x'] = loaded_params.get('right_x')
            result['up_y'] = loaded_params.get('up_y')
            result['down_y'] = loaded_params.get('down_y')
            # Добавляем настройки PNA и синхронизатора
            result['pna_settings'] = loaded_params.get('pna_settings')
            result['sync_settings'] = loaded_params.get('sync_settings')
        
        return result
        
    except Exception as e:
        logger.error(f"Ошибка при загрузке результатов измерения лучей: {e}", exc_info=True)
        return None
