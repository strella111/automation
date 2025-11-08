from typing import Tuple, List, Dict
import os
import datetime

from loguru import logger
from openpyxl import Workbook

from utils.excel_module import get_or_create_excel_for_check
from config.settings_manager import get_main_settings
from ...devices.afar import Afar
from ...devices.pna import PNA
from core.devices.trigger_box import E5818
from core.common.enums import Channel, Direction
from ...common.exceptions import WrongInstrumentError
import threading
from PyQt5.QtCore import QThread


class CheckAfarStend:
    """Класс для проверки АФАР на стенде с обратным триггером"""

    def __init__(self, afar: Afar,
                 pna: PNA,
                 gen: E5818,
                 bu_numbers: List[int],  # Список номеров БУ для проверки
                 stop_event: threading.Event = None,
                 pause_event: threading.Event = None,
                 check_fv: bool = True,  # Проверять ФВ
                 check_lz: bool = True):  # Проверять ЛЗ
        """
        Инициализация класса

        Args:
            afar: АФАР
            pna: Анализатор цепей
            gen: Устройство синхронизации (TriggerBox)
            bu_numbers: Список номеров БУ (МА) для проверки
            stop_event: Событие для остановки измерений
            pause_event: Событие для приостановки измерений
            check_fv: Проверять фазовращатели
            check_lz: Проверять линии задержки

        Raises:
            ValueError: Если какой-либо из параметров None
        """
        if not all([afar, pna, gen, bu_numbers, stop_event, pause_event]):
            raise ValueError("Все устройства, список БУ и события должны быть указаны")

        self.afar = afar
        self.pna = pna
        self.gen = gen
        self.bu_numbers = bu_numbers
        self._stop_event = stop_event or threading.Event()
        self._pause_event = pause_event or threading.Event()
        self.check_fv = check_fv
        self.check_lz = check_lz

        self.period = None
        self.number_of_freqs = None
        self.lead = None

        self.phase_shifts = [0, 5.625, 11.25, 22.5, 45, 90, 180]
        self.delay_lines = [0, 1, 2, 4, 8]

        self.channel = None
        self.direction = None

        self.norm_amp = None
        self.norm_phase = None
        self.norm_delay = None

        self.delay_amp_tolerance = 1.0
        self.delay_tolerances = {
            1: {'min': 90.0, 'max': 110.0},
            2: {'min': 180.0, 'max': 220.0},
            4: {'min': 360.0, 'max': 440.0},
            8: {'min': 650, 'max': 800}
        }

        self.delay_callback = None
        self.data_callback = None
        self.realtime_callback = None
        self.amp_data_callback = None  # Колбэк для данных только амплитуды
        self.bu_completed_callback = None  # Колбэк для уведомления о завершении измерения БУ

        self.max_ppm_count = 32
        self.max_phase_shifts = len(self.phase_shifts)
        self.max_delay_lines = len(self.delay_lines)

        self.data_real = None
        self.data_relative = None

    def _clear_memory(self):
        """Очистка данных из памяти для предотвращения накопления"""
        if self.data_real is not None:
            self.data_real.clear()
            self.data_real = None
        if self.data_relative is not None:
            self.data_relative.clear()
            self.data_relative = None
        logger.debug("Память очищена от данных измерений")

    def _get_memory_usage_info(self) -> str:
        """Получение информации об использовании памяти"""
        import sys
        
        real_size = sys.getsizeof(self.data_real) if self.data_real else 0
        relative_size = sys.getsizeof(self.data_relative) if self.data_relative else 0

        real_items = sum(len(v) for v in self.data_real.values()) if self.data_real else 0
        relative_items = sum(len(v) for v in self.data_relative.values()) if self.data_relative else 0
        
        return f"Память: real={real_size} байт ({real_items} элементов), relative={relative_size} байт ({relative_items} элементов)"

    def _check_connections(self) -> bool:
        """
        Проверка подключения всех устройств

        Returns:
            bool: True если все устройства подключены, False в противном случае
        """
        if not all([self.pna.connection, self.afar.connection, self.gen.connection]):
            logger.error("Не все устройства подключены")
            return False
        return True

    def burst_and_check_external_trigger(self, bu_num: int, ppm_num: int):
        self.gen.burst(period_s=self.period, count=self.number_of_freqs, lead_s=self.lead)
        QThread.msleep(int((self.lead + self.period*self.number_of_freqs)*1000))
        counter = 0
        while True:
            evt = self.gen.pop_ext_event()
            counter += 1
            if counter > 5:
                try:
                    tm_data = self.afar.get_tm(bu_num=bu_num)
                    if tm_data is None:
                        logger.warning(f"Не удалось получить телеметрию для БУ {bu_num}, ППМ {ppm_num}")
                        return

                    amount_strobs = tm_data['strobs_prd'] if self.channel == Channel.Transmitter else tm_data['strobs_prm']
                    expected_strobs = (ppm_num - 1) * self.number_of_freqs
                    if amount_strobs == expected_strobs:
                        self.gen.burst(period_s=self.period, count=self.number_of_freqs, lead_s=self.lead)
                        while True:
                            evt = self.gen.pop_ext_event()
                            if evt:
                                return
                    else:
                        return
                except Exception as e:
                    logger.error(f"Ошибка при проверке телеметрии для БУ {bu_num}, ППМ {ppm_num}: {e}")
                    return

            if evt:
                return

    def _normalize_phase(self, phase: float) -> float:
        """Нормализует фазу в диапазон [-180, 180]"""
        while phase > 180:
            phase -= 360
        while phase < -180:
            phase += 360
        return phase

    def _check_fv(self, bu_num: int, chanel: Channel, direction: Direction) -> Dict[float, List[float]]:
        """Проверка фазовращателей для конкретного БУ"""
        results: Dict[float, List[float]] = {}
        zero_phases: List[float] = []

        # Ограничиваем размер данных для предотвращения накопления памяти
        max_data_per_fv = self.max_ppm_count * 2  # 2 значения на ППМ (амплитуда, фаза)
        for fv in self.phase_shifts:
            code = int(fv / 5.625)
            self.afar.set_calb_mode(bu_num=bu_num,
                                  chanel=chanel,
                                  direction=direction,
                                  delay_number=0,
                                  fv_number=code,
                                  att_ppm_number=0,
                                  att_mdo_number=0,
                                  number_of_strobes=self.number_of_freqs)

            if fv not in results:
                results[fv] = []
            ppm_index = 0

            for ppm_num in range(1, 33):
                if self._stop_event.is_set():
                    self.pna.set_output(False)
                    logger.info("Измерение остановлено пользователем")
                    return results

                if self._pause_event.is_set():
                    logger.info("Измерение приостановлено пользователем")

                while self._pause_event.is_set() and not self._stop_event.is_set():
                    QThread.msleep(100)

                self.burst_and_check_external_trigger(bu_num=bu_num, ppm_num=ppm_num)

                amp_db, phase_deg = self.pna.get_center_freq_data()
                

                if len(results[fv]) < max_data_per_fv:
                    results[fv].extend([amp_db, phase_deg])
                else:
                    logger.warning(f"Превышен лимит данных для ФВ={fv}, пропускаем ППМ {ppm_num}")


                if fv == 0:
                    if len(zero_phases) < self.max_ppm_count:
                        zero_phases.append(phase_deg)
                    phase_rel = 0.0
                else:
                    phase_zero = zero_phases[ppm_index] if ppm_index < len(zero_phases) else 0.0
                    phase_rel = self._normalize_phase(phase_zero - phase_deg)
                    if phase_rel < 0:
                        phase_rel += 360

                if self.realtime_callback:
                    try:
                        self.realtime_callback.emit(float(fv), int(ppm_index + 1), float(amp_db), float(phase_rel), int(bu_num))
                    except Exception as e:
                        logger.error(f"Ошибка realtime обновления UI: {e}")

                ppm_index += 1
        return results

    def _check_amp_only(self, bu_num: int, chanel: Channel, direction: Direction) -> List[float]:
        """Проверка только амплитуды ППМ (без циклов по ФВ и ЛЗ)"""
        # Устанавливаем режим калибровки с ФВ=0, ЛЗ=0
        self.afar.set_calb_mode(bu_num=bu_num,
                              chanel=chanel,
                              direction=direction,
                              delay_number=0,
                              fv_number=0,
                              att_ppm_number=0,
                              att_mdo_number=0,
                              number_of_strobes=self.number_of_freqs)
        
        amp_data = []
        
        for ppm_num in range(1, 33):
            if self._stop_event.is_set():
                self.pna.set_output(False)
                logger.info("Измерение остановлено пользователем")
                return amp_data

            if self._pause_event.is_set():
                logger.info("Измерение приостановлено пользователем")
            while self._pause_event.is_set() and not self._stop_event.is_set():
                QThread.msleep(100)

            self.burst_and_check_external_trigger(bu_num=bu_num, ppm_num=ppm_num)

            try:
                amp_db, _ = self.pna.get_center_freq_data()
            except Exception:
                amp_db = float('nan')

            if len(amp_data) < self.max_ppm_count:
                amp_data.append(amp_db)
            else:
                logger.warning(f"Превышен лимит данных для амплитуды, пропускаем ППМ {ppm_num}")

            if self.amp_data_callback:
                try:
                    self.amp_data_callback.emit(amp_data.copy(), int(bu_num))
                except Exception as e:
                    logger.error(f"Ошибка realtime обновления UI (амплитуда): {e}")
        
        return amp_data

    def _check_lz(self, bu_num: int, chanel: Channel, direction: Direction) -> Dict[int, tuple]:
        """Проверка линий задержки для конкретного БУ"""
        self.pna.set_delay_type()

        lz_to_amps: Dict[int, List[float]] = {}
        lz_to_delays: Dict[int, List[float]] = {}
        zero_amp_mean: float = None
        zero_delay_mean: float = None

        for lz in self.delay_lines:
            lz_to_amps[lz] = []
            lz_to_delays[lz] = []

            self.afar.set_calb_mode(bu_num=bu_num,
                                  chanel=chanel,
                                  direction=direction,
                                  delay_number=lz,
                                  fv_number=0,
                                  att_ppm_number=0,
                                  att_mdo_number=0,
                                  number_of_strobes=self.number_of_freqs)

            collected = 0
            for ppm_num in range(1, 33):
                if self._stop_event.is_set():
                    self.pna.set_output(False)
                    logger.info("Измерение остановлено пользователем")
                    return {}

                if self._pause_event.is_set():
                    logger.info("Измерение приостановлено пользователем")
                while self._pause_event.is_set() and not self._stop_event.is_set():
                    QThread.msleep(100)

                self.burst_and_check_external_trigger(bu_num=bu_num, ppm_num=ppm_num)

                try:
                    delay_mean = float(self.pna.get_mean_value_from_fdata()) * 1e12
                except Exception:
                    delay_mean = float('nan')

                try:
                    amp_mean = float(self.pna.get_mean_value_from_sdata())
                except Exception:
                    amp_mean = float('nan')

                # Проверяем размер данных перед добавлением
                if len(lz_to_amps[lz]) < self.max_ppm_count:
                    lz_to_amps[lz].append(amp_mean)
                    lz_to_delays[lz].append(delay_mean)
                    collected += 1
                else:
                    logger.warning(f"Превышен лимит данных для ЛЗ={lz}, пропускаем ППМ {ppm_num}")

            mean_amp = (sum(lz_to_amps[lz]) / len(lz_to_amps[lz])) if lz_to_amps[lz] else float('nan')
            mean_delay = (sum(lz_to_delays[lz]) / len(lz_to_delays[lz])) if lz_to_delays[lz] else float('nan')

            if lz == 0:
                zero_amp_mean = mean_amp
                zero_delay_mean = mean_delay
                amp_delta_rt = 0.0
                delay_delta_rt = 0.0
            else:
                if zero_amp_mean is None or zero_delay_mean is None:
                    amp_delta_rt = mean_amp
                    delay_delta_rt = mean_delay
                else:
                    amp_delta_rt = (mean_amp - zero_amp_mean)
                    delay_delta_rt = (mean_delay - zero_delay_mean)

            if self.delay_callback:
                try:
                    self.delay_callback.emit({lz: (amp_delta_rt, delay_delta_rt)}, bu_num)
                except Exception as e:
                    logger.error(f"Ошибка realtime обновления таблицы ЛЗ: {e}")

        results: Dict[int, tuple] = {}
        zero_amp = None
        zero_delay = None
        if 0 in lz_to_amps and len(lz_to_amps[0]) > 0:
            zero_amp = sum(lz_to_amps[0]) / len(lz_to_amps[0])
        if 0 in lz_to_delays and len(lz_to_delays[0]) > 0:
            zero_delay = sum(lz_to_delays[0]) / len(lz_to_delays[0])

        for lz in self.delay_lines:
            amps = lz_to_amps.get(lz, [])
            delays = lz_to_delays.get(lz, [])
            mean_amp = (sum(amps) / len(amps)) if amps else float('nan')
            mean_delay = (sum(delays) / len(delays)) if delays else float('nan')

            if zero_amp is not None and zero_delay is not None:
                amp_delta = mean_amp - zero_amp
                delay_delta = mean_delay - zero_delay
            else:
                amp_delta = mean_amp
                delay_delta = mean_delay

            if lz == 0:
                amp_delta = 0.0
                delay_delta = 0.0

            results[lz] = (amp_delta, delay_delta)

        return results

    def start(self, channel: Channel, direction: Direction):
        """
        Запуск проверки всех выбранных БУ

        Returns:
            List[Tuple[int, Tuple[bool, Tuple[float, float]]]]:
            Список кортежей (номер БУ, результаты проверки)

        Raises:
            ConnectionError: При отсутствии подключения устройств
            WrongInstrumentError: При ошибке работы с устройствами
        """
        results = []
        
        try:
            if not self._check_connections():
                raise ConnectionError("Не все устройства подключены")

            self.pna.set_output(True)
            self.pna.set_ascii_data()

            try:
                self.number_of_freqs = int(self.pna.get_amount_of_points())
            except Exception as e:
                logger.error(f"Не удалось получить количество точек с PNA: {e}")
                raise

            # Для режима только амплитуды создаем один файл для всех БУ
            amp_only_workbook = None
            amp_only_worksheet = None
            amp_only_file_path = None
            amp_only_row_counter = 1  # Счетчик строк для записи данных
            
            if not self.check_fv and not self.check_lz:

                channel_str = 'ПРМ' if channel == Channel.Receiver else 'ПРД'
                direction_str = 'Г' if direction == Direction.Horizontal else 'В'

                date_str = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                file_name = f"{date_str}_{channel_str}{direction_str}.xlsx"

                base_dir = get_main_settings().value('base_save_dir', '')
                dir_path = os.path.join(base_dir, 'check_afar', 'amp')
                if not os.path.exists(dir_path):
                    os.makedirs(dir_path)
                
                amp_only_file_path = os.path.join(dir_path, file_name)

                if os.path.exists(amp_only_file_path):
                    from openpyxl import load_workbook
                    amp_only_workbook = load_workbook(amp_only_file_path)
                    amp_only_worksheet = amp_only_workbook.active
                    amp_only_row_counter = amp_only_worksheet.max_row + 1
                else:
                    amp_only_workbook = Workbook()
                    amp_only_worksheet = amp_only_workbook.active
                    amp_only_worksheet.cell(1, 1).value = 'Номер БУ'
                    amp_only_worksheet.cell(1, 2).value = 'Номер ППМ'
                    amp_only_worksheet.cell(1, 3).value = 'Канал'
                    amp_only_worksheet.cell(1, 4).value = 'Поляризация'
                    amp_only_worksheet.cell(1, 5).value = 'Амплитуда (дБ)'
                    amp_only_row_counter = 2

            for bu_num in self.bu_numbers:
                if self._stop_event.is_set():
                    logger.info("Измерение остановлено пользователем")
                    break
                
                logger.info(f"Начало проверки БУ №{bu_num}")

                if not self.check_fv and not self.check_lz:
                    # Только амплитуда
                    check_mode = 'amp'
                    subdir = 'amp'
                elif self.check_fv and not self.check_lz:
                    check_mode = 'amp_fv'
                    subdir = 'amp_fv'
                elif self.check_fv and self.check_lz:
                    check_mode = 'amp_fv_lz'
                    subdir = 'amp_fv_lz'
                else:
                    check_mode = 'amp_lz'
                    subdir = 'amp_lz'

                worksheet = None
                workbook = None
                file_path = None
                
                if self.check_fv or self.check_lz:
                    worksheet, workbook, file_path = get_or_create_excel_for_check(
                        base_dir=get_main_settings().value('base_save_dir', ''),
                        dir_name=os.path.join('check_afar', subdir),
                        file_name=f'БУ{bu_num}.xlsx',
                        mode='stend',
                        chanel=channel,
                        direction=direction)

                try:
                    if not self.check_fv and not self.check_lz:
                        amp_data = self._check_amp_only(bu_num=bu_num, chanel=channel, direction=direction)

                        if amp_only_worksheet is not None:
                            channel_str = 'Приемник' if channel == Channel.Receiver else 'Передатчик'
                            direction_str = 'Горизонтальная' if direction == Direction.Horizontal else 'Вертикальная'
                            
                            for ppm_num in range(1, 33):
                                row = amp_only_row_counter
                                amp_only_worksheet.cell(row, 1).value = bu_num
                                amp_only_worksheet.cell(row, 2).value = ppm_num
                                amp_only_worksheet.cell(row, 3).value = channel_str
                                amp_only_worksheet.cell(row, 4).value = direction_str
                                
                                if ppm_num - 1 < len(amp_data):
                                    amp_only_worksheet.cell(row, 5).value = amp_data[ppm_num - 1]
                                else:
                                    amp_only_worksheet.cell(row, 5).value = 0.0
                                
                                amp_only_row_counter += 1

                            try:
                                amp_only_workbook.save(amp_only_file_path)
                                logger.debug(f"Данные БУ №{bu_num} сохранены в общий файл")
                            except Exception as e:
                                logger.error(f"Ошибка промежуточного сохранения общего файла для БУ {bu_num}: {e}")

                    if self.check_fv:
                        data = self._check_fv(bu_num=bu_num, chanel=channel, direction=direction)
                        self.data_real = data
                        logger.debug(f"После _check_fv для БУ {bu_num}: {self._get_memory_usage_info()}")

                        if 0 in data:
                            rel_data: Dict[float, List[float]] = {}
                            zero_list = data[0]
                            max_relative_data = self.max_ppm_count * 2  # Ограничиваем размер относительных данных
                            
                            for fv, values in data.items():
                                rel_list: List[float] = []
                                for i in range(0, min(len(values), max_relative_data), 2):
                                    if i + 1 < len(values):
                                        amp_val = values[i]
                                        phase_val = values[i + 1]
                                        phase_zero = zero_list[i + 1] if i + 1 < len(zero_list) else 0.0
                                        phase_rel = self._normalize_phase(phase_zero - phase_val) if fv != 0 else 0.0
                                        rel_list.extend([amp_val, phase_rel])
                                rel_data[fv] = rel_list
                            self.data_relative = rel_data
                            logger.debug(f"После обработки относительных данных для БУ {bu_num}: {self._get_memory_usage_info()}")
                        else:
                            logger.warning(f'Не найдены данные для ФВ=0 для БУ {bu_num}. Относительные фазы не будут сформированы')
                            self.data_relative = None

                        if self.data_callback and self.data_relative:
                            try:
                                self.data_callback.emit(self.data_relative, int(bu_num))
                            except Exception as e:
                                logger.error(f"Ошибка отправки данных ФВ в UI: {e}")

                        if self.data_relative:
                            for ppm_num in range(1, 33):
                                excel_row = [ppm_num, ]
                                for fv, res in self.data_relative.items():
                                    amp_index = (ppm_num-1) * 2
                                    phase_index = (ppm_num-1) * 2 + 1
                                    
                                    if amp_index < len(res) and phase_index < len(res):
                                        excel_row.append(res[amp_index])
                                        excel_row.append(res[phase_index])
                                    else:
                                        logger.warning(f"Недостаточно данных для ППМ {ppm_num} в ФВ {fv} для БУ {bu_num}")
                                        excel_row.extend([0.0, 0.0])  # Заполняем нулями при отсутствии данных
                                        
                                for k, value in enumerate(excel_row):
                                    worksheet.cell(row=ppm_num + 2, column=k + 1).value = value

                    if self.check_lz:
                        try:
                            lz_results = self._check_lz(bu_num=bu_num, chanel=channel, direction=direction)

                            excel_row = [f'ЛЗ№', 'Относительная амплитуда', 'Задержка, пс']
                            for i, value in enumerate(excel_row):
                                worksheet.cell(row=36, column=i+1).value = value
                            for lz, result in lz_results.items():
                                if lz == 0:
                                    continue
                                excel_row = [f'ЛЗ{lz}', result[0], result[1]]
                                for i, value in enumerate(excel_row):
                                    worksheet.cell(row=36 + self.delay_lines.index(lz), column=i + 1).value = value

                        except Exception as e:
                            logger.error(f"Ошибка проверки линий задержки для БУ {bu_num}: {e}")

                    if workbook is not None and file_path is not None:
                        workbook.save(file_path)
                    logger.info(f"Проверка БУ №{bu_num} завершена")

                    if self.bu_completed_callback:
                        try:
                            self.bu_completed_callback.emit(int(bu_num))
                        except Exception as e:
                            logger.error(f"Ошибка уведомления о завершении БУ {bu_num}: {e}")

                    
                except Exception as e:
                    logger.error(f"Ошибка при проверке БУ {bu_num}: {e}")
                finally:
                    pass

            try:
                self.pna.set_output(False)
            except Exception as e:
                logger.error(f"Ошибка при выключении PNA: {e}")
                raise WrongInstrumentError(f"Ошибка выключения PNA: {e}")

            if amp_only_workbook is not None and amp_only_file_path is not None:
                try:
                    amp_only_workbook.save(amp_only_file_path)
                    logger.info(f"Данные всех БУ сохранены в общий файл: {amp_only_file_path}")
                except Exception as e:
                    logger.error(f"Ошибка сохранения общего файла для режима только амплитуды: {e}")

            self._clear_memory()
            logger.info("Измерение всех БУ завершено")
            
        except Exception as e:
            logger.error(f"Ошибка при выполнении проверки: {e}")
            try:
                self.pna.set_output(False)
            except Exception as e:
                logger.error(f"Ошибка при аварийном выключении PNA: {e}")

            self._clear_memory()
            raise

        return results

