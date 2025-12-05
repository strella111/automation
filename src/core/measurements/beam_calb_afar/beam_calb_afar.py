from typing import List
import os
import datetime

from loguru import logger
from openpyxl import Workbook

from config.settings_manager import get_main_settings
from ...devices.afar import Afar
from ...devices.pna import PNA
from core.devices.trigger_box import E5818
from core.common.enums import Channel, Direction
import threading
from PyQt5.QtCore import QThread


class BeamAfarCalb:

    def __init__(self, afar: Afar,
                 pna: PNA,
                 gen: E5818,
                 bu_numbers: List[int],
                 beam_numbers: List[int],
                 with_calb: bool = True,
                 stop_event: threading.Event = None,
                 pause_event: threading.Event = None):

        """
        Инициализация класса

        Args:
                        afar: АФАР
            pna: Анализатор цепей
            gen: Устройство синхронизации (TriggerBox)
            bu_numbers: Список номеров БУ (МА) для проверки
            stop_event: Событие для остановки измерений
            pause_event: Событие для приостановки измерений

        """
        # if not all([afar, pna, gen, bu_numbers, stop_event, pause_event]):
        #     raise ValueError("Все устройства, список БУ и события должны быть указаны")

        self.afar = afar
        self.pna = pna
        self.gen = gen
        self.bu_numbers = bu_numbers
        self.beam_numbers = beam_numbers
        self.with_calb = with_calb
        self._stop_event = stop_event or threading.Event()
        self._pause_event = pause_event or threading.Event()

        self.period = None
        self.number_of_freqs = None
        self.lead = None

        self.chanel = None
        self.direction = None

        self.data = None


    def _check_connections(self) -> bool:
        """
        Проверка подключения всех устройств

        Returns:
            bool: True если все устройства подключены, False в противном случае
        """
        if not all([self.pna.connection, self.afar.connection, self.gen.connection]):
            logger.error("Не все устройства подключены")
            return False
        return True

    def burst_and_check_external_trigger(self, bu_num: int, ppm_num: int):
        self.gen.burst(period_s=self.period, count=self.number_of_freqs, lead_s=self.lead)
        QThread.msleep(int((self.lead + self.period*self.number_of_freqs)*1000))
        counter = 0
        while True:
            evt = self.gen.pop_ext_event()
            counter += 1
            if counter > 5:
                try:
                    tm_data = self.afar.get_tm(bu_num=bu_num)
                    if tm_data is None:
                        logger.warning(f"Не удалось получить телеметрию для БУ {bu_num}, ППМ {ppm_num}")
                        return

                    amount_strobs = tm_data['strobs_prd'] if self.channel == Channel.Transmitter else tm_data['strobs_prm']
                    expected_strobs = (ppm_num - 1) * self.number_of_freqs
                    if amount_strobs == expected_strobs:
                        self.gen.burst(period_s=self.period, count=self.number_of_freqs, lead_s=self.lead)
                        while True:
                            evt = self.gen.pop_ext_event()
                            if evt:
                                return
                    else:
                        return
                except Exception as e:
                    logger.error(f"Ошибка при проверке телеметрии для БУ {bu_num}, ППМ {ppm_num}: {e}")
                    return

            if evt:
                return


    def _normalize_phase(self, phase: float) -> float:
        """Нормализует фазу в диапазон [-180, 180]"""
        while phase > 180:
            phase -= 360
        while phase < -180:
            phase += 360
        return phase

    def start(self, chanel: Channel,
              direction: Direction,
              beams: List[int],
              freq_list: List[float]):

        logger.info(f'Начало измерения лучей через калибровочный канал. Лучи {beams}. Частоты {len(freq_list)}')

        if not self.data:
            self.data = {}

        channel_str = 'ПРМ' if chanel == Channel.Receiver else 'ПРД'
        direction_str = 'Г' if direction == Direction.Horizontal else 'В'

        date_str = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

        base_dir = get_main_settings().value('base_save_dir', '')
        dir_path = os.path.join(base_dir, 'beams', 'calb_beams', channel_str + direction_str + '_' + date_str)
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)


        try:
        #     if not self._check_connections():
        #         raise ConnectionError('Не все устройства подключены')

            self.pna.set_output(True)
            self.pna.set_ascii_data()

            try:
                self.number_of_freqs = int(self.pna.get_amount_of_points())
            except Exception as e:
                logger.error(f"Не удалось получить количество точек с PNA: {e}")
                raise


            for beam in self.beam_numbers:
                file_name = f'beam№{beam}.xlsx'
                file_path = os.path.join(dir_path, file_name)

                workbook = Workbook()
                worksheet = workbook.active
                worksheet.cell(1, 1).value = 'Номер БУ'
                worksheet.cell(1, 2).value = 'Номер ППМ'
                worksheet.cell(1, 3).value = 'Канал'
                worksheet.cell(1, 4).value = 'Амплитуда (дБ)'
                worksheet.cell(1, 5).value = 'Фаза'

                count = -1
                for bu_num in self.bu_numbers:
                    count += 1
                    if not self.data.get(bu_num):
                        self.data[bu_num] = {}

                    if self._stop_event.is_set():
                        logger.info("Измерение остановлено пользователем")
                        break

                    if self._pause_event.is_set():
                        QThread.msleep(100)

                    self.afar.preset_task(bu_num=bu_num)

                    self.afar.set_beam_calb_mode(bu_num=bu_num,
                                                 beam_number=beam,
                                                 chanel=chanel,
                                                 direction=direction,
                                                 with_calb=self.with_calb,
                                                 table_num=1,
                                                 table_crc=b'\x00\x00',
                                                 amount_strobs=self.number_of_freqs)
                    for ppm_num in range(1, 33):
                        #self.burst_and_check_external_trigger(bu_num=bu_num, ppm_num=ppm_num)

                        amp, phase = self.pna.get_center_freq_data()

                        if not self.data[bu_num].get(beam):
                            self.data[bu_num][beam] = []
                        self.data[bu_num][beam].append((amp, phase))

                        row = ppm_num + count * 32 + 1
                        worksheet.cell(row, 1).value = bu_num
                        worksheet.cell(row, 2).value = ppm_num
                        worksheet.cell(row, 3).value = channel_str + direction_str
                        worksheet.cell(row, 4).value = amp
                        worksheet.cell(row, 5).value = phase

                    workbook.save(file_path)


        except Exception as e:
            logger.error(f"Ошибка при выполнении измерения лучей: {e}")

            try:
                self.pna.set_output(False)
            except Exception as e:
                logger.error(f"Ошибка при аварийном выключении PNA: {e}")
            self.data = None
            raise









